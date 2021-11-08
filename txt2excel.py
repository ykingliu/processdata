# coding: utf-8
# import xlwt
import openpyxl

# xlrd 和 xlwt 是python中用来处理 xls 文件的函数，其单个 sheet 限制最大行数为65535
# openpyxl 函数，其最大行数为1048576，存储的文件类型为 xlsx

txtname = 'F:\\20211026\\mono\\20001-\\20001-.txt'
excelname = 'F:\\20211026\\mono\\20001-\\20001-.xlsx'


def writetoxlsx():
    # data = open('/home/liuyuanqing/downloads/demo1/8-15s.txt', 'r', encoding='UTF-8')
    data = open(txtname, 'r', encoding='UTF-8')
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    lines = data.readlines()

    i = 1  # 注意：'cell'函数中行列起始值为1
    for line in lines:
        line = line.strip('\n')
        line = line.split(' ')
        for x in range(0, 2):
            # 对应音频名不带.wav
            # outws.cell(column = x+1 , row = i , value = "%s" % line[x]+".wav" if x==0 else line[x])
            # 对应音频名带有.wav
            outws.cell(column=x + 1, row=i, value="%s" % line[x] if x == 0 else line[x])
            # outws.cell(column = x+1 , row = i , value = "%s" % line[x])
        i += 1
    savexlsx = "/home/liuyuanqing/downloads/demo1/8-15s.xlsx"
    outwb.save(excelname)  # 保存结果
    data.close()

writetoxlsx()
